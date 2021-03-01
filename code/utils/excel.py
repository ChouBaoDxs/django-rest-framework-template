from datetime import datetime
from io import BytesIO
import typing

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet, Cell
import xlwt

# 自动换行样式
wrap_style = xlwt.XFStyle()
wrap_style.alignment.wrap = 1


def generate_excel_io(excel_data: typing.List,
                      excel_header: typing.List = None,
                      sheet_name: str = '') -> BytesIO:
    sheet_name = sheet_name or '1'

    output = BytesIO()
    work_book = xlwt.Workbook(encoding="utf-8")
    work_sheet = work_book.add_sheet(sheet_name)

    offset = 0
    # 处理表头
    if excel_header:
        for index, title in enumerate(excel_header):
            work_sheet.write(0, index, title)
        offset = 1

    # 处理内容
    for row_num, row_data in enumerate(excel_data):
        for index, cell_data in enumerate(row_data):
            work_sheet.write(row_num + offset, index, cell_data, wrap_style)

    work_book.save(output)
    output.seek(0)

    return output


def generate_excel_io_by_openpyxl(excel_data: typing.List,
                                  excel_header: typing.List = None,
                                  sheet_name: str = '') -> BytesIO:
    sheet_name = sheet_name or '1'

    wb: openpyxl.Workbook = openpyxl.Workbook()
    # ws: Worksheet = wb.active
    # 删除默认的sheet
    wb.remove(wb.get_active_sheet())

    ws: Worksheet = wb.create_sheet(sheet_name)
    # openpyxl 的行和列从 1 开始

    offset = 0
    # 处理表头
    if excel_header:
        for index, title in enumerate(excel_header):
            # 第一种写入方式：
            # one_cell = ws.cell(1, index)
            # one_cell.value = data

            # 第二种写入方式：
            ws.cell(1, index + 1, value=title)
        offset = 1

    for row_num, row_data in enumerate(excel_data):
        for index, cell_data in enumerate(row_data):
            ws.cell(1 + row_num + offset, index + 1, value=cell_data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_xlsx_by_openpyxl(excel_data: typing.List,
                              excel_header: typing.List = None,
                              sheet_name: str = '',
                              save_file_path=''):
    save_file_path = save_file_path or '{}.xlsx'.format(datetime.now().strftime("%Y%m%d%H%M%S"))
    content = generate_excel_io_by_openpyxl(excel_data, excel_header, sheet_name)
    with open(save_file_path, 'wb') as f:
        f.write(content.read())
